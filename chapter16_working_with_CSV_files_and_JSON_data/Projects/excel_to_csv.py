# excel_to_csv.py 
# This program scans the current folder, opens each Excel file, reads every sheet,
# and writes the cell values into separate CSV files automatically.


import os
import openpyxl
import csv



for excel_file in os.listdir('.'):
    # Skip non-xlsx files, load the workbook object.
    if not excel_file.endswith('.xlsx'):
        continue

    print(f'Opening {excel_file}...')

    wb = openpyxl.load_workbook(excel_file)

    # Loop through every sheet in the workbook.
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Create the CSV filename from the excel filename and sheet title
        csv_file_name = f'{excel_file[:-5]}_{sheet_name}.csv'
        print(f'wrriting {csv_file_name}....')

        # Create the CSV.writer object for this CSV file.
        csv_new = open(csv_file_name, 'w', newline='')
        csv_writer = csv.writer(csv_new)

        # Loop through every row in the sheet.
        for row_num in range(1, sheet.max_row + 1):
            row_data = [] # Append each cell to this list

            # Loop through each cell in the row.
            for col_num in range(1, sheet.max_column + 1):
                
                # Append each cell's data tp row_data.
                obj_text = sheet.cell(row=row_num, column=col_num).value
                row_data.append(obj_text)
        
            # Write the row_data list to the CSV file
            csv_writer.writerow(row_data)

        csv_new.close()
    wb.close()