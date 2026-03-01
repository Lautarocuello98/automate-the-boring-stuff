# Lautarocuello98
# multiplication_table.py

import sys
import openpyxl
from openpyxl.styles import Font


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: py multiplication_table.py <Number>")

    try:
        n = int(sys.argv[1])
        if n < 1:
            raise ValueError
    except ValueError:
        raise SystemExit("The number must be a positive integer.")

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Multiplication Table"

    bold = Font(bold=True)

    # Header row/column
    for i in range(1, n + 1):
        sheet.cell(row=1, column=i + 1, value=i).font = bold
        sheet.cell(row=i + 1, column=1, value=i).font = bold

    # Table values
    for row in range(1, n + 1):
        for col in range(1, n + 1):
            sheet.cell(row=row + 1, column=col + 1, value=row * col)

    # Nice-to-have UX
    sheet.freeze_panes = "B2"
    sheet.column_dimensions["A"].width = 6
    for col in range(2, n + 2):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 6

    filename = f"MultiplicationTable_{n}x{n}.xlsx"
    wb.save(filename)
    print(f"✅ Created {filename}")


if __name__ == "__main__":
    main()