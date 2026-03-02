# Lautarocuello98
# spreadsheet_to_text.py
# Writes each column of the active sheet into a separate .txt file.

from pathlib import Path
import openpyxl

# Get Excel file from user
xlsx = Path(input("Enter the spreadsheet name: ").strip().strip('"'))
if not xlsx.exists():
    raise SystemExit(f"File not found: {xlsx}")

# Create output folder
out = Path("output")
out.mkdir(exist_ok=True)

# Load workbook and select active sheet
wb = openpyxl.load_workbook(xlsx)
sheet = wb.active

files = []
for col in range(1, sheet.max_column + 1):

    # Use header as filename (fallback if empty)
    name = sheet.cell(1, col).value
    name = str(name).strip() if name else f"column_{col}"

    # Sanitize filename (avoid invalid characters)
    safe = "".join(c for c in name if c.isalnum() or c in (" ", "_", "-")).rstrip()
    txt = out / f"{safe}.txt"

    # Write column values to file
    with txt.open("w", encoding="utf-8") as f:
        for row in range(1, sheet.max_row + 1):
            v = sheet.cell(row, col).value
            if v is not None:
                f.write(f"{v}\n")

    files.append(txt.name)

print(f"Done ✅ Created {len(files)} files in {out}/")