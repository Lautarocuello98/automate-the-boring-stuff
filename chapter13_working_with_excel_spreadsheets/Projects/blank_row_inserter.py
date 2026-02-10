import sys
import sys
import openpyxl

# Validate the number of command-line arguments.
# The script expects:
# 1) N → the starting row where blank rows will be inserted
# 2) M → how many blank rows to insert
# 3) filename → Excel file to modify
if len(sys.argv) != 4:
    print("Usage: python blank_row_inserter.py <N> <M> filename.xlsx")
    sys.exit(1)

# Read and convert arguments
n = int(sys.argv[1])      # Row number where insertion begins
m = int(sys.argv[2])      # Number of blank rows to insert
filename = sys.argv[3]    # Excel file name

# Open the workbook and select the active worksheet
wb = openpyxl.load_workbook(filename)
sheet = wb.active

# Move existing data down by M rows.
# We iterate from the bottom upward to avoid overwriting data.
for row in range(sheet.max_row, n - 1, -1):
    for col in range(1, sheet.max_column + 1):
        # copy the value of each cell to its new position
        sheet.cell(row=row + m, column=col).value = \
            sheet.cell(row=row, column=col).value
        

# Clear the original cells to create the blank rows.
# This ensures rows N to N+M-1 remain empty.
for row in range(n, n + m):
    for col in range(1, sheet.max_column + 1):
        sheet.cell(row=row, column=col).value = None




# Save the modified workbook (overwrites the original file)
wb.save(filename)

# Confirmation message
print(f"Inserted {m} blank rows starting at row {n} in {filename}")