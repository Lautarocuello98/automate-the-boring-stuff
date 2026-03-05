# text_to_spreadsheet.py - Lautarocuello98
# Read multiple text files and write their contents into an Excel spreadsheet.
# Each file becomes a column, each line becomes a row.

from pathlib import Path
import openpyxl

def main():
    # Collect file names from the user
    filenames = []
    while True:
        name = input("Give me a file (Enter to finish): ").strip()
        if not name:
            break
        filenames.append(name)

    if not filenames:
        print("No files provided.")
        return

    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Text to Spreadsheet"

    # Process each file and write its contents into a separate column
    for col, filename in enumerate(filenames, start=1):
        path = Path(filename)

        # Write the file name as the column header
        sheet.cell(row=1, column=col).value = path.name

        try:
            # Open the text file and write each line into the spreadsheet
            with path.open(encoding="utf-8") as f:
                for row, line in enumerate(f, start=2):  # start below header
                    sheet.cell(row=row, column=col).value = line.strip()

        except FileNotFoundError:
            print(f"File not found: {filename}, skipping.")
            sheet.cell(row=2, column=col).value = "(file not found)"

    # Save the workbook to disk
    wb.save("output.xlsx")
    print("Done. Spreadsheet saved as output.xlsx")

if __name__ == "__main__":
    main()
    