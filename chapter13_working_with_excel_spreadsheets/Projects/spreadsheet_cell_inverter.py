import openpyxl

# open file
filename = input("what is the file?: ").strip
wb = openpyxl.load_workbook(filename)
sheet = wb.active

# create a new sheet for the inverted data
new_sheet = wb.create_sheet(title="Inverted")

# copy cells inverted (transpose)
for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        new_sheet.cell(row=col, column=row).value = sheet.cell(row=row, column=col).value

# save result
wb.save(f"inverted_{filename}")