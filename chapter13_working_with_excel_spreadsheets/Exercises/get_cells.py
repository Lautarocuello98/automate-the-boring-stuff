import openpyxl

wb = openpyxl.load_workbook('example.xlsx')

# get a sheet from the workbook
sheet = wb['Sheet1']
# get a cell from the sheet and the value
print(sheet['A1'].value)

# get another cell from the sheet
c = sheet['B1']
print(c.value)

# get the row, column, and value from the cell.
print(f'Row {c.row}, Column {c.column} is: {c.value}')

print(f'cell {c.coordinate} is {c.value}')

print(sheet['C1'].value)

# go through every other row:
print()
for i in range(1, 8, 2):
    print(f'{i} {sheet.cell(row=i, column=2).value}')

# get the highest row number.
print(sheet.max_row)

# get the highest column number.
print(sheet.max_column)
